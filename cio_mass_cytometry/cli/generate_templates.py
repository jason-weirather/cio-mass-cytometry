""" Generate templates of inputs for cio-mass-cytometry.
Will create a form for filling in the panel and samples.

""" 

import os
from datetime import datetime
import argparse
import json
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from importlib_resources import files
from cio_mass_cytometry.utilities import get_validator

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True)
highlight.fill = PatternFill(start_color="CDEAC2", end_color="CDEAC2", fill_type = "solid")

annotation_highlight = NamedStyle(name="annotation_highlight")
annotation_highlight.font = Font(bold=True)
annotation_highlight.fill = PatternFill(start_color="ffb6c1", end_color="ffb6c1", fill_type = "solid")

boldened = NamedStyle(name="boldened")
boldened.font = Font(bold=True)

def cmd_create(args):
   # If we create, make a new file at the template path
   if os.path.exists(args.template_path) and args.overwrite is False:
      raise ValueError("Cannot overwrite without --overwrite option")
   create_template(args.template_path)
def cmd_validate(args):
   # If we valdiate, its's easymode. Just read it
   print("validate")
def cmd_add_samples(args):
   # If we add to it, we modify it.
   print("add samples")
def cmd_ingest(args):
   # Read in and write out json
   print("ingest")

def create_template(template_path):
   _validator1 = get_validator(files('schemas').joinpath('panel.json'))
   _schema1 = _validator1.schema

   _validator2 = get_validator(files('schemas').joinpath('files.json'))
   _schema2 = _validator2.schema

   wb = Workbook()
   default_names = wb.sheetnames
   wb.add_named_style(highlight)


   # Start with the Metadata. Write the header and the value names
   ws0 = wb.create_sheet("Panel Parameters")
   _write_parameters(ws0,[_schema1['properties']['parameters']])
   _fix_width(ws0)

   ws1 = wb.create_sheet("Panel Definition")
   _write_repeating(ws1,_schema1['properties']['markers'])
   _fix_width(ws1)


   example_annotations = ["Batch","Timepoint","Response"]
   ws2 = wb.create_sheet("Sample Manifest")
   ncols = _write_repeating(ws2,_schema2['properties']['samples'],exclude=['sample_annotations'])
   _write_annotations(ws2,example_annotations,start_idx = ncols)
   _fix_width(ws2)

   ws3 = wb.create_sheet("Sample Annotations")
   _write_repeating(ws3,_schema2['definitions']['sample_annotations'])
   _fix_width(ws3)

   # cleanup workbook deleting default sheet name
   for _sheet_name in default_names:
      del wb[_sheet_name]
   wb.save(filename = template_path)
   return

def _write_parameters(worksheet,fields):
   "Write the metadata fields to the worksheet"
   "fields can be either a single object of metadata, or multiple in list or tuple to stack"
   header_names = ['Parameter','Value']
   for _j,_header_name in enumerate(header_names):
      worksheet.cell(row=1,column=_j+1).style = highlight
      worksheet.cell(row=1,column=_j+1).value = _header_name 
   
   is_list = False
   if type(fields) is list:
      is_list = True
   elif type(fields) is tuple:
      is_list = True

   # If we got a list of fields consoldate their properties
   if is_list:
      _f = {}
      _f['properties'] = {}
      for _field in fields:
         for _k in _field['properties']:
            _f['properties'][_k] = _field['properties'][_k]
      fields = _f

   for _i, _property in enumerate(fields['properties']):
      worksheet.cell(row=_i+2,column=1).style = boldened
      worksheet.cell(row=_i+2,column=1).value=  fields['properties'][_property]['title']
      if 'default' in fields['properties'][_property]:
         worksheet.cell(row=_i+2,column=2).value = fields['properties'][_property]['default']
         
         

def _write_repeating(worksheet,fields,exclude=[]):
   "Write the repeating data fields to the worksheet"
   header_names = list(fields['items']['properties'])
   #print(header_names)
   ncols = 0
   for _j,_header_name in enumerate(header_names):
      print(_header_name)
      if _header_name in exclude: continue

      entry = fields['items']['properties'][_header_name]
      worksheet.cell(row=1,column=_j+1).style = highlight
      worksheet.cell(row=1,column=_j+1).value = _header_name if 'title' not in entry else entry['title']
      ncols+=1
   return ncols

def _write_annotations(worksheet,fields,start_idx=0):
   for _j, _header_name in enumerate(fields):
      worksheet.cell(row=1,column=_j+1+start_idx).style = annotation_highlight
      worksheet.cell(row=1,column=_j+1+start_idx).value = _header_name 
   return

def _fix_width(worksheet,min_width=20,padding=3):
   column_widths = []
   for row in worksheet:
      for i, cell in enumerate(row):
         if cell.value is None: continue
         if len(column_widths) > i:
            if len(str(cell.value)) > column_widths[i]:
                column_widths[i] = min_width if len(str(cell.value))+padding < min_width else len(str(cell.value))+padding
         else:
            column_widths += [min_width if len(str(cell.value))+padding < min_width else len(str(cell.value))+padding]
   for i, column_width in enumerate(column_widths):
      worksheet.column_dimensions[get_column_letter(i+1)].width = column_width


def do_report_output(output_path):
   _validator = get_validator(files('schema_data.inputs').joinpath('report_definition.json'))
   _schema = _validator.schema
   wb = Workbook()
   default_names = wb.sheetnames
   wb.add_named_style(highlight)


   # Start with the Metadata. Write the header and the value names

   ws0 = wb.create_sheet(_schema['properties']['parameters']['title'])
   _write_parameters(ws0,_schema['properties']['parameters'])
   _fix_width(ws0)

   ws1 = wb.create_sheet(_schema['properties']['population_percentages']['title'])
   _write_repeating(ws1,_schema['properties']['population_percentages'])
   _fix_width(ws1)

   ws2 = wb.create_sheet(_schema['properties']['population_densities']['title'])
   _write_repeating(ws2,_schema['properties']['population_densities'])
   _fix_width(ws2)

   # cleanup workbook deleting default sheet name
   for _sheet_name in default_names:
      #print(_sheet_name)
      del wb[_sheet_name]
   wb.save(filename = output_path)
   return



def do_project_folder_output(output_file):
   # For now lets keep this with InForm only
   _validator = get_validator(files('schema_data.inputs.platforms.InForm').joinpath('project.json'))
   _schema = _validator.schema

   wb = Workbook()
   default_names = wb.sheetnames
   wb.add_named_style(highlight)

   # Start with the Metadata. Write the header and the value names

   ws1 = wb.create_sheet(_schema['properties']['parameters']['title'])
   _write_parameters(ws1,_schema['properties']['parameters'])
   _fix_width(ws1)


   # Now lets make the Panel.  Write the header only.
   ws2 = wb.create_sheet(_schema['properties']['samples']['title'])
   _write_repeating(ws2,_schema['properties']['samples'])
   _fix_width(ws2)

   # cleanup workbook deleting default sheet name
   for _sheet_name in default_names:
      #print(_sheet_name)
      del wb[_sheet_name]
   wb.save(filename = output_file)

def do_panel_output(args):
   #import schema_data.inputs as schema_data_inputs

   _validator = get_validator(files('schema_data.inputs').joinpath('panel.json'))
   _schema = _validator.schema

   wb = Workbook()
   default_names = wb.sheetnames
   wb.add_named_style(highlight)

   _oname = args.panel_output

   # Start with the Metadata. Write the header and the value names

   ws1 = wb.create_sheet(_schema['properties']['parameters']['title'])
   _write_parameters(ws1,_schema['properties']['parameters'])
   _fix_width(ws1)


   # Now lets make the Panel.  Write the header only.
   ws2 = wb.create_sheet(_schema['properties']['markers']['title'])
   _write_repeating(ws2,_schema['properties']['markers'])
   _fix_width(ws2)

   # cleanup workbook deleting default sheet name
   for _sheet_name in default_names:
      #print(_sheet_name)
      del wb[_sheet_name]
   wb.save(filename = _oname)

def cli():
   parser = argparse.ArgumentParser(
            description = "",
            formatter_class=argparse.ArgumentDefaultsHelpFormatter)
   subparsers = parser.add_subparsers(help='Choose how to work with the tempate')

   parser_create = subparsers.add_parser('create', help='Create a new template')
   parser_create.add_argument("--overwrite",action="store_true",help="Allow file overwriting on creation")
   parser_create.set_defaults(func=cmd_create)

   parser_add = subparsers.add_parser('add_samples', help='Add samples to a template.')
   parser_add.add_argument("--sample_path",help="Path where samples are stored")
   parser_add.set_defaults(func=cmd_add_samples)

   parser_validate = subparsers.add_parser('validate', help='Validate an existing template.')
   parser_validate.set_defaults(func=cmd_validate)

   parser_ingest = subparsers.add_parser('ingest', help='Read an existing template.')
   parser_ingest.set_defaults(func=cmd_ingest)

   parser.add_argument('template_path',help="Path to write the template")

   args = parser.parse_args()
   args.func(args)
   return 


if __name__ == "__main__":
	cli()