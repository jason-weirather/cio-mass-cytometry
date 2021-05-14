import logging, os, re
from openpyxl import load_workbook

logging.basicConfig(level=logging.WARN)
logger = logging.getLogger()

fcs_path_column = 4 # 1-indexed column number
fcs_column_label = 'FCS File'
sample_name_column = 1 # 1-indexed column number
sample_name_label = 'Sample Name'

def add_samples(template_path,samples_path, output_path, sample_name_regex, mylogger):
    logger = mylogger
    if sample_name_regex:
        logger.info("Regex present.  Compile the regex.")
        sample_name_regex = re.compile(sample_name_regex)

    wb = load_workbook(template_path)
    ws = wb['Sample Manifest']
    insert_row, current_fcs_files = get_lowest_row(ws)
    # Now lets get the sample list
    fcs_list = find_paths(samples_path)
    # See if there are overlaps
    overlap = set(fcs_list).intersection(set(current_fcs_files))
    if len(overlap) > 0:
        raise ValueError("The following "+str(len(overlap))+" files were already present in the sheet.\n"+"\n".join(sorted(overlap)))
    # Add the samples to the excel sheet
    for i, fname in enumerate(fcs_list):
        # Add the sample name
        ws.cell(row=insert_row+i+1,column=fcs_path_column).value = fname
        # If we have a regex and a match, add the sample name
        if sample_name_regex:
            m = sample_name_regex.search(fname)
            if not m and not m.group(1):
                logger.warn("Failed to match group regex for (row "+str(insert_row+i+1)+") "+fname)
                continue
            ws.cell(row=insert_row+i+1,column=sample_name_column).value = m.group(1)

    wb.save(output_path)
def find_paths(samples_path):
    fcs_list = []
    for root, dirs, fnames in os.walk(samples_path):
        for fname in fnames:
            if fname[-4:] != '.fcs': continue
            fcs_list.append(os.path.join(root,fname))
    return fcs_list

def get_lowest_row(worksheet):
    lowest = -1
    current_fcs_files = []
    for i, row in enumerate(worksheet.iter_rows(max_row=0)):
        if len([x for x in row if x.value])>0: lowest=i
        if i==0:
            if row[fcs_path_column-1].value != fcs_column_label: raise ValueError("Must add to '"+fcs_column_label+"' column") 
            if row[sample_name_column-1].value != sample_name_label: raise ValueError("Ill defined '"+sample_name_label+"' column") 
            continue
        if row[fcs_path_column-1].value: 
            current_fcs_files.append(row[fcs_path_column-1].value)
    if lowest==-1: 
        raise ValueError("Failed to find a bottom row.")
    logger.info("Finished finidng lowest row "+str(lowest)+" with "+str(len(current_fcs_files))+" fcs files already prsent.")
    return lowest+1, current_fcs_files