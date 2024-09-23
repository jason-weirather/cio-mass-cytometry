""" Generate templates of inputs for cio-mass-cytometry.
Will create a form for filling in the panel and samples.
""" 

import os, json, logging
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from importlib_resources import files
from cio_mass_cytometry.utilities import get_validator, get_version
from cio_mass_cytometry import schemas


highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True)
highlight.fill = PatternFill(start_color="CDEAC2", end_color="CDEAC2", fill_type = "solid")

annotation_highlight = NamedStyle(name="annotation_highlight")
annotation_highlight.font = Font(bold=True)
annotation_highlight.fill = PatternFill(start_color="ffb6c1", end_color="ffb6c1", fill_type = "solid")

boldened = NamedStyle(name="boldened")
boldened.font = Font(bold=True)

logging.basicConfig(level=logging.WARN)
logger = logging.getLogger()


def create_template(template_path,mylogger):
   logger = mylogger
   logger.info("Reading and validating panel.json")
   _validator1 = get_validator(files(schemas).joinpath('panel.json'))
   _schema1 = _validator1.schema

   logger.info("Reading and validating samples.json")
   _validator2 = get_validator(files(schemas).joinpath('samples.json'))
   _schema2 = _validator2.schema

   logger.info("Reading and validating pipeline.json")
   _validator3 = get_validator(files(schemas).joinpath('pipeline.json'))
   _schema3 = _validator3.schema

   logger.info("Creating the Workbook")
   wb = Workbook()
   default_names = wb.sheetnames
   wb.add_named_style(highlight)


   logger.info("Start with Panel Metadata")
   ws0 = wb.create_sheet("Panel Parameters")
   _write_parameters(ws0,[_schema1['properties']['parameters']])
   _fix_width(ws0)

   logger.info("Set the panel definition")
   ws1 = wb.create_sheet("Panel Definition")
   _write_repeating(ws1,_schema1['properties']['markers'])
   _fix_width(ws1)


   logger.info("Create the sample manifest and add some example annotation columns")
   example_annotations = ["Batch","Timepoint","Response"]
   ws2 = wb.create_sheet("Sample Manifest")
   ncols = _write_repeating(ws2,_schema2['properties']['samples'],exclude=['sample_annotations'])
   _write_annotations(ws2,example_annotations,start_idx = ncols)
   _fix_width(ws2)

   logger.info("Make a table to define the sample annotations")
   ws3 = wb.create_sheet("Sample Annotations")
   _write_repeating(ws3,_schema2['properties']['annotation_levels'])
   _fix_width(ws3)

   logger.info("End with the pipeline version")
   ws4 = wb.create_sheet("Meta")
   _write_parameters(ws4,[_schema3])
   _fix_width(ws4)

   # Manually set the version
   ws4.cell(row=2,column=2).value = get_version() 

   # cleanup workbook deleting default sheet name
   for _sheet_name in default_names:
      del wb[_sheet_name]
   wb.save(filename = template_path)
   return

def _write_parameters(worksheet,fields):
   logger.info("Write the metadata fields to the worksheet")
   header_names = ['Parameter','Value']
   for _j,_header_name in enumerate(header_names):
      worksheet.cell(row=1,column=_j+1).style = highlight
      worksheet.cell(row=1,column=_j+1).value = _header_name 
   
   is_list = False
   if type(fields) is list:
      is_list = True
   elif type(fields) is tuple:
      is_list = True

   # If we got a list of fields consoldate their properties
   if is_list:
      _f = {}
      _f['properties'] = {}
      for _field in fields:
         for _k in _field['properties']:
            _f['properties'][_k] = _field['properties'][_k]
      fields = _f

   for _i, _property in enumerate(fields['properties']):
      worksheet.cell(row=_i+2,column=1).style = boldened
      worksheet.cell(row=_i+2,column=1).value=  fields['properties'][_property]['title']
      if 'default' in fields['properties'][_property]:
         worksheet.cell(row=_i+2,column=2).value = fields['properties'][_property]['default']

def _write_repeating(worksheet,fields,exclude=[]):
   logger.info("Write the repeating data fields to the worksheet")
   header_names = list(fields['items']['properties'])
   #print(header_names)
   ncols = 0
   for _j,_header_name in enumerate(header_names):
      if _header_name in exclude: continue

      entry = fields['items']['properties'][_header_name]
      worksheet.cell(row=1,column=_j+1).style = highlight
      worksheet.cell(row=1,column=_j+1).value = _header_name if 'title' not in entry else entry['title']
      ncols+=1
   return ncols

def _write_annotations(worksheet,fields,start_idx=0):
   for _j, _header_name in enumerate(fields):
      worksheet.cell(row=1,column=_j+1+start_idx).style = annotation_highlight
      worksheet.cell(row=1,column=_j+1+start_idx).value = _header_name 
   return

def _fix_width(worksheet,min_width=20,padding=3):
   column_widths = []
   for row in worksheet:
      for i, cell in enumerate(row):
         if cell.value is None: continue
         if len(column_widths) > i:
            if len(str(cell.value)) > column_widths[i]:
                column_widths[i] = min_width if len(str(cell.value))+padding < min_width else len(str(cell.value))+padding
         else:
            column_widths += [min_width if len(str(cell.value))+padding < min_width else len(str(cell.value))+padding]
   for i, column_width in enumerate(column_widths):
      worksheet.column_dimensions[get_column_letter(i+1)].width = column_width
